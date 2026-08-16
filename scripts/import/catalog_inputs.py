#!/usr/bin/env python3
"""Discover and parse Input A (monthly coupons) and Input B (referrals).

Implements section 2 (Input A/B contracts) and section 3 (workspace
discovery and sample validation) of the master prompt. Standard library
only, except optional XLSX support via ``openpyxl`` when it is installed.
"""

from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None

COUPON_HEADERS = {
    "course_id", "course_name", "coupon_type", "maximum_redemptions",
    "coupon_code", "start_date_time", "end_date_time", "currency",
    "discount_price", "course_coupon_url",
}
REFERRAL_REQUIRED_HEADERS = {"course_id", "course_name"}
REFERRAL_URL_HEADER_ALIASES = (
    "Referal URL for students",
    "Referral URL for students",
)

SUPPORTED_OFFER_TYPES = {"best_price", "custom_price", "free_open", "free_targeted"}
OFFER_TYPE_ALIASES = {"current_best_price": "best_price"}

EXCLUDED_DIR_NAMES = {
    ".git", ".github", "node_modules", "__pycache__", "data", "assets",
    "scripts", "tests", "docs", "templates", "config", ".vscode",
}


@dataclass
class RowError:
    row_index: int
    reason: str

    def to_dict(self) -> dict[str, Any]:
        return {"rowIndex": self.row_index, "reason": self.reason}


@dataclass
class ParsedTable:
    source_file: str
    sha256: str
    kind: str  # "coupons" | "referrals"
    rows: list[dict[str, str]] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.errors


def sha256_of(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_xlsx_sheets(path: Path) -> dict[str, list[dict[str, str]]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read %s" % path.name)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, str]]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
        except StopIteration:
            sheets[sheet_name] = []
            continue
        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if raw_row is None or all(cell is None for cell in raw_row):
                continue
            row = {}
            for key, value in zip(header, raw_row):
                if not key:
                    continue
                row[key] = "" if value is None else str(value)
            rows.append(row)
        sheets[sheet_name] = rows
    return sheets


def classify_headers(headers: Iterable[str]) -> str | None:
    header_set = set(headers)
    if COUPON_HEADERS.issubset(header_set):
        return "coupons"
    has_referral_url = any(alias in header_set for alias in REFERRAL_URL_HEADER_ALIASES)
    if REFERRAL_REQUIRED_HEADERS.issubset(header_set) and has_referral_url:
        return "referrals"
    return None


def discover_candidate_tables(root: Path) -> list[tuple[Path, str, str]]:
    """Return (path, sheet_name_or_'', kind) for every CSV/XLSX candidate found."""
    candidates: list[tuple[Path, str, str]] = []
    for path in sorted(root.iterdir()):
        if path.is_dir() or path.name.startswith("."):
            continue
        suffix = path.suffix.lower()
        if suffix == ".csv":
            try:
                rows = _read_csv_rows(path)
            except (OSError, csv.Error):
                continue
            if not rows:
                continue
            kind = classify_headers(rows[0].keys())
            if kind:
                candidates.append((path, "", kind))
        elif suffix == ".xlsx":
            try:
                sheets = _read_xlsx_sheets(path)
            except (RuntimeError, OSError):
                continue
            for sheet_name, rows in sheets.items():
                if not rows:
                    continue
                kind = classify_headers(rows[0].keys())
                if kind:
                    candidates.append((path, sheet_name, kind))
    return candidates


def dedupe_by_hash(candidates: list[tuple[Path, str, str]]) -> tuple[list[tuple[Path, str, str]], list[str]]:
    """Process one copy of byte-identical duplicates; report the rest."""
    seen_hashes: dict[str, Path] = {}
    unique: list[tuple[Path, str, str]] = []
    skipped: list[str] = []
    for path, sheet_name, kind in candidates:
        digest = sha256_of(path)
        if digest in seen_hashes:
            skipped.append(
                "%s (sheet '%s') is byte-identical to %s and was skipped."
                % (path.name, sheet_name, seen_hashes[digest].name)
            )
            continue
        seen_hashes[digest] = path
        unique.append((path, sheet_name, kind))
    return unique, skipped


def _row_reader(path: Path, sheet_name: str) -> list[dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx_sheets(path)[sheet_name]
    return _read_csv_rows(path)


def _normalize_start_end(row: dict[str, str], index: int, errors: list[RowError]) -> None:
    for field_name in ("start_date_time", "end_date_time"):
        value = (row.get(field_name) or "").strip()
        if not value:
            errors.append(RowError(index, "%s is missing." % field_name))


def parse_coupons(path: Path, sheet_name: str = "") -> ParsedTable:
    raw_rows = _row_reader(path, sheet_name)
    errors: list[RowError] = []
    rows: list[dict[str, str]] = []
    seen_ids: set[str] = set()

    for index, raw_row in enumerate(raw_rows, start=2):
        row = {key: (value or "").strip() for key, value in raw_row.items()}
        course_id = row.get("course_id", "")
        if not course_id:
            errors.append(RowError(index, "course_id is missing."))
            continue
        if course_id in seen_ids:
            errors.append(RowError(index, "Duplicate course_id '%s' in coupon export." % course_id))
            continue
        seen_ids.add(course_id)

        offer_type_raw = row.get("coupon_type", "")
        offer_type = OFFER_TYPE_ALIASES.get(offer_type_raw, offer_type_raw)
        if offer_type not in SUPPORTED_OFFER_TYPES:
            errors.append(RowError(index, "Unsupported coupon_type '%s'." % offer_type_raw))
            continue

        missing = [key for key in ("course_name", "course_coupon_url") if not row.get(key)]
        if missing:
            errors.append(RowError(index, "Missing required field(s): %s." % ", ".join(missing)))
            continue
        if not row["course_coupon_url"].startswith("https://"):
            errors.append(RowError(index, "course_coupon_url is not an HTTPS URL."))
            continue

        _normalize_start_end(row, index, errors)

        row["course_id"] = course_id
        row["coupon_type"] = offer_type
        rows.append(row)

    return ParsedTable(source_file=str(path), sha256=sha256_of(path), kind="coupons", rows=rows, errors=errors)


def parse_referrals(path: Path, sheet_name: str = "") -> ParsedTable:
    raw_rows = _row_reader(path, sheet_name)
    errors: list[RowError] = []
    rows: list[dict[str, str]] = []
    seen_ids: set[str] = set()

    for index, raw_row in enumerate(raw_rows, start=2):
        row = {key: (value or "").strip() for key, value in raw_row.items()}
        course_id = row.get("course_id", "")
        if not course_id:
            errors.append(RowError(index, "course_id is missing."))
            continue
        if course_id in seen_ids:
            errors.append(RowError(index, "Duplicate course_id '%s' in referral sheet." % course_id))
            continue
        seen_ids.add(course_id)

        referral_url = ""
        for alias in REFERRAL_URL_HEADER_ALIASES:
            if row.get(alias):
                referral_url = row[alias]
                break
        if not referral_url:
            errors.append(RowError(index, "Referral URL column is missing or empty."))
            continue
        if not referral_url.startswith("https://www.udemy.com/") and not referral_url.startswith("https://udemy.com/"):
            errors.append(RowError(index, "instructor_referral_url is not an HTTPS Udemy URL."))
            continue
        if "referralCode=" not in referral_url:
            errors.append(RowError(index, "instructor_referral_url has no referralCode query parameter."))
            continue

        rows.append(
            {
                "course_id": course_id,
                "course_name": row.get("course_name", ""),
                "instructor_referral_url": referral_url,
            }
        )

    return ParsedTable(source_file=str(path), sha256=sha256_of(path), kind="referrals", rows=rows, errors=errors)


def parse_iso_offer_datetime(value: str) -> datetime | None:
    """Parse Udemy's 'YYYY-MM-DD HH:MM TZ' export format (PST/PDT aware)."""
    value = value.strip()
    if not value:
        return None
    tz_offsets = {
        "PST": -8 * 60, "PDT": -7 * 60,
        "EST": -5 * 60, "EDT": -4 * 60,
        "UTC": 0, "GMT": 0,
    }
    parts = value.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].upper() in tz_offsets:
        naive_part, tz_name = parts
        offset_minutes = tz_offsets[parts[1].upper()]
        try:
            naive = datetime.strptime(naive_part, "%Y-%m-%d %H:%M")
        except ValueError:
            return None
        from datetime import timedelta, timezone

        return naive.replace(tzinfo=timezone(timedelta(minutes=offset_minutes)))
    for pattern in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            from datetime import timezone

            parsed = datetime.strptime(value, pattern)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None
